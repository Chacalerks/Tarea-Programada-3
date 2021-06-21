from openpyxl import Workbook
import openpyxl
import os
import webbrowser

def abrirFile(nombreFile):
    """
    funcionamiento: se encarga de abrir el archivo que sea indicado por el nombre con la aplicacion por defecto
    entradas: nombreFile: el nombre del archivo a abrir
    salidas: N/A
    """
    os.startfile(nombreFile)

def abrirPage(nombreFile):
    """
    funcionamiento: se encarga de abrir el archivo que sea indicado por el nombre con el navegador
    entradas: nombreFile: el nombre del archivo a abrir
    salidas: N/A
    """
    webbrowser.open_new_tab(nombreFile)


def reporteFichaLarga(nombreArch):
    """
    Funcionamiento: crea el archivo excel con la ficha mas larga que incluye todos los datos de una licencia
    Entradas: nombreArch: el nombre con el que se va a guardar el archivo
    Salidas: NA
    """
    lista = ['Licencia LicenciaA2','LicenciaLicencia A3','Licencia LicenciaB1','LicenciaLicencia B2','Licencia LicenciaB3','Licencia LicenciaB4',
'LicenciaLicencia C1','Licencia LicenciaC2','LicenciaLicencia D1','LicenciaLicencia D2','LicenciaLicencia D3','LicenciaLicencia E1','Licencia LicenciaE2']
    workbook = Workbook()
    sheet = workbook.active

    sheet['A3'] = 'Cédula'
    sheet['B3'] = 'Nombre'
    sheet['C3'] = 'FechaNac'
    sheet['D3'] = 'FechaExp'
    sheet['E3'] = 'FechaVenc'
    sheet['F3'] = 'TipoLicen'
    sheet['G3'] = 'TipoSangre'
    sheet['H3'] = 'Donador'
    sheet['I3'] = 'Sede'
    sheet['J3'] = 'Puntaje'
    cont = 2
    for i in lista:
        sheet['A'+str(cont)] = i
        cont+=1
    
    workbook.save(nombreArch)

reporteFichaLarga("ferks.xlsx")
abrirPage("ferks.xlsx")
input("Picherks")
wb = openpyxl.load_workbook("ferks.xlsx")
worksheet = wb.active
for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        if cell.coordinate in worksheet.merged_cells: # not check merge_cells
            continue
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column].width = adjusted_width
wb.save("ferks.xlsx")
abrirPage("ferks.xlsx")
input("Picherks")