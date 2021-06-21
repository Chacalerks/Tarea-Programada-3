#Creado por: César Jiménez Salazar y Maynor Erks Martínez Hernández.
#Fecha de realización:05/06/2021 07:21 p.m
#Última modificación:21/06/2021 10:59  a.m
#Versión: 3.9.5

from validaciones import *
import names
import random
from datetime import datetime
from openpyxl import Workbook
from dateutil.relativedelta import relativedelta
import time
from dominate.tags import *
from archivo import *
from clase import *
import os
import webbrowser
import openpyxl
"""
DOCUMENTACIÓN

+ La estrategia para generar las sedes y ligarlas a una licencia, fue crear una lista donde el código que se almacenará
de esa sede en la base de datos, será la posición que ocupa en dicha lista, y los nombres de cada sede se decidieron por lo
más característico de la localización de cada una:
Sede Central
0 = 'San José, San Sebastián'

Licencias en el Gran Área Metropolitana
1 = 'Montecillos Alajuela'
2 = 'Tránsito Cartago'
3 = 'Barva de Heredia'
4 = 'Tránsito San Ramón'

Licencias en el Atlántico
5 = 'Guapiles, Ruta 32'
6 = 'Barrio Sandoval de Moín'

Licencias en Guanacaste y Puntarenas
7 = 'Carretera al Aeropuerto Daniel Oduber'
8 = 'Aeropuerto de Nicoya'
9 = 'Chacarita, Calle 138'

Licencias en la Zona Sur
10 = 'Pérez Zeledón'
11 = 'Río Claro de Golfito'

Licencias en la Zona Norte
12 = 'San Carlos'
"""
#----------------------------------------------------------------------------#
#                           Funciones Generales                              #
#----------------------------------------------------------------------------#
def determinarPar(num):
    """
    Funcionamiento: Determina si el número es par
    Entradas:
    -num(int): número a válidar
    Salidas: 
    -True: si es par
    -False: si no es par
    """
    if num % 2 ==0:
        return True
    else:
        return False

def obtenerHoraActual():
    """
    funcionamiento: Se encarga de obtner la hora actual
    entradas: Na
    salidas: Hora Actual
    """
    return datetime.now().time().strftime('%H:%M:%S')

def convertirFechaAnnos(fecha):
    """
    funcionamiento: convierte una fecha de nacimiento a los annos que debe tener esa persona
    entradas: fecha: la fecha de nacimiento
    salidas: los annos que debe tener la persona que nacio en esa fecha
    """
    annos = int(generarFechaExp()[6:]) - int(fecha[6:])
    return annos

def obtenerPosicion(cedula, lista):
    """
    funcionamiento: retorna la posicion del objeto en la lista con la cedula especificada
    entradas: cedula: la cedula a buscar
    salidas: la posicion del objeto en la lista con la cedula especificada 
    """
    for i in lista:
        if cedula == i.getCedula():
            return lista.index(i)
    return False

def obtenerPath():
    """
    funcionamiento: se encarga de devolver en string la direccion del archivo que se está ejecutando en cualquier computador
    entradas: N/A
    salidas: N/A
    """
    return os.path.dirname(os.path.abspath(__file__))

def abrirPage(nombreFile):
    """
    funcionamiento: se encarga de abrir el archivo que sea indicado por el nombre con el navegador
    entradas: nombreFile: el nombre del archivo a abrir
    salidas: N/A
    """
    webbrowser.open_new_tab(nombreFile)

def abrirFile(nombreFile, carpeta):
    """
    funcionamiento: se encarga de abrir el archivo que sea indicado por el nombre con la aplicacion por defecto
    entradas: nombreFile: el nombre del archivo a abrir
    salidas: N/A
    """
    os.startfile(obtenerPath()+"/"+carpeta+"//"+nombreFile)

#----------------------------------------------------------------------------#
#                           Generar Licencia                                 #
#----------------------------------------------------------------------------#

def str_time_prop(start, end, time_format, prop):
    """
    funcionamiento: genera una fecha aleatoria entre dos fechas dadas
    entradas: start: la fecha inicial
    end: la fecha futura
    time_format: el formato de la fecha
    prop: numero aleatorio para sacar la fecha
    salidas: la fecha aleatoria
    """
    stime = time.mktime(time.strptime(start, time_format))
    etime = time.mktime(time.strptime(end, time_format))

    ptime = stime + prop * (etime - stime)

    return time.strftime(time_format, time.localtime(ptime))

def randomDate(start, end, prop):
    """
    funcionamiento: Una fecha randon
    entradas: start: la fecha inicial   -end: la fecha futura -prop: numero aleatorio para sacar la fecha
    salidas: la fecha aleatoria
    """
    return str_time_prop(start, end, '%d-%m-%Y', prop)

def generarRangoFecha():
    """
    funcionamiento: genera el rango correcto de la fecha
    entradas: NA
    salidas:    Un rango de fecha
    """
    fechaMin = (datetime.today() + relativedelta(years=-18)).strftime('%d-%m-%Y')
    fechaMax = (datetime.today() + relativedelta(years=-50)).strftime('%d-%m-%Y')
    return [fechaMax,fechaMin]
    

def generarLicencias(cant, tiposLicencias, lista):
    """
    funcionamiento: genera la cantidad de lincencias aleatorias
    entradas: cant, tiposLicencias, lista
    salidas: Guarda las licencias aleatorias
    """
    cant = int(cant)
    while cant != 0:
        insertarLicencia(randomLicencia(tiposLicencias),lista)
        cant-=1
    
def randomLicencia(tiposLicencias):
    """
    funcionamiento: genera una licencia con informacion aleatoria
    entradas: tiposLicencias: los subtipos de licencias
    salidas: el objeto licencia con los datos
    """
    licencia = [generarCedula(), generarNombre(), generarFechaNacimiento(), generarFechaExp()]
    licencia.append(generarFechaVenc(convertirFechaAnnos(licencia[2])))
    licencia.append(generarTipoLicencia(tiposLicencias))
    licencia.append(generarTipoSangre())
    licencia.append(generarDonador())
    licencia.append(generarSede(licencia[0][0]))
    licencia.append(generarPuntaje())
    licencia.append(generarCorreo(licencia[1]))
    licencia1 = Licencia(licencia[0],licencia[1],licencia[2],licencia[3],licencia[4],licencia[5],licencia[6],licencia[7],licencia[8],
                         licencia[9], licencia[10])
    return licencia1
    
def generarCedula():
    """
    funcionamiento: genera un numero de cedula aleatorio
    entradas: N/A
    salidas: un numero de cedula aleatorio
    """
    cedula = str(random.randint(1,9))+str(random.randint(1000,9999))+str(random.randint(1000,9999))
    return cedula

def generarNombre():
    """
    funcionamiento: genera un nombre aleatorio 
    entradas: N/A
    salidas: un nombre aleatorio
    """
    nombre = names.get_full_name()+" "+names.get_last_name()
    return nombre

def generarFechaNacimiento():
    """
    funcionamiento: genera una fecha de nacimiento aleatoria
    entradas: N/A
    salidas: un nombre aleatorio
    """
    dob = randomDate(generarRangoFecha()[0], generarRangoFecha()[1], random.random())

    if validarFormatoFecha(dob):
        return dob
    else:
        return generarFechaNacimiento()

def generarFechaExp():
    """
    funcionamiento: genera la fecha de hoy
    entradas: N/A
    salidas: la fecha de hoy en string
    """
    fecha = time.strftime('%d-%m-%Y', time.localtime())

    if validarFormatoFecha(fecha):
        return fecha
    else:
        return generarFechaExp()

def generarFechaVenc(edad):
    """
    funcionamiento: genera la fecha de vencimiento de la licencia dependiendo de su edad
    entradas: la edad de la persona aplicando para la licencia
    salidas: la fecha de vencimiento
    """
    if edad >= 18 and edad <= 25:
        edadVen = generarFechaExp()[0:6] + str(int(generarFechaExp()[6:])+3)
    else:
        edadVen = generarFechaExp()[0:6] + str(int(generarFechaExp()[6:])+5)

    if validarFormatoFecha(edadVen):
        return edadVen
    else:
        return generarFechaVenc(edad)

def generarTipoLicencia(tiposLicencias):
    """
    funcionamiento: retorna un tipo de licencia aleatorio
    entradas: la lista con los tipos de licencias
    salidas: el tipo escogido de manera aleatoria
    """
    return random.choice(tiposLicencias)

def generarTipoSangre():
    """
    funcionamiento: retorna un tipo de sangre aleatorio
    entradas: N/A
    salidas: el tipo escogido de manera aleatoria
    """
    return random.choice(["O", "A", "B", "AB"])+random.choice("-+")

def generarDonador():
    """
    funcionamiento: retorna un valor booleano aleatorio
    entradas: N/A
    salidas: el tipo escogido de manera aleatoria
    """
    return random.choice([True, False])

def generarSede(id):
    """
    funcionamiento: retorna una sede aleatoria con respecto a la cedula
    entradas: id: la inicial de la cedula
    salidas: la sede escogida
    """
    if id in ['1','8','9']:
        return [random.choice(['0','10'])]
    elif id == '2':
        return [random.choice(['1','4','12'])]
    elif id == '3':
        return ['2']
    elif id == '4':
        return ['3']
    elif id == '5':
        return [random.choice(['7','8'])]
    elif id == '6':
        return [random.choice(['9','11'])]
    else:
        return [random.choice(['5','6'])]

def generarPuntaje():
    """
    funcionamiento: genera un numero entre 0 y 12, que representa el puntaje
    entradas: N/A
    salidas: el puntaje
    """
    puntaje = random.randint(0,12)
    if validarPuntajeFormato(puntaje):
        return puntaje
    else:
        return generarPuntaje()


def generarCorreo(nombre):
    """
    funcionamiento: genera un correo electrónico basado en el nombre de la persona
    entradas: nombre: el nombre de la persona
    salidas: el correo
    """
    partes = nombre.split()
    correo = (partes[1]+partes[2][0]+partes[0][0]).lower() + "@gmail.com"
    if validarCorreo(nombre, correo):
        return correo
    else:
        return generarCorreo(nombre)

#----------------------------------------------------------------------------#
#                           Traducciones                                     #
#----------------------------------------------------------------------------#

#Las funciones de traducir se encargan de hacer la conversion de un valor almacenado de cierta manera
# en la base de datos, y devolver su equivalente mostrable al usuario y viceversa

def traducirSede(sede):
    """
    funcionamiento: Traduce el sede según corresponda
    entradas: sede: sede a traducir
    salidas: Traduccion del sede
    """
    sedes = ['San José, San Sebastián','Montecillos Alajuela','Tránsito Cartago','Barva de Heredia','Tránsito San Ramón','Guapiles, Ruta 32',
               'Barrio Sandoval de Moín','Carretera al Aeropuerto Daniel Oduber','Aeropuerto de Nicoya','Chacarita, Calle 138','Pérez Zeledón',
               'Río Claro de Golfito','San Carlos']
    try:
        if sede.isdigit():
            return sedes[int(sede)]
        else:
            return str(sedes.index(sede))
    except:
        return sede

def traducirSedeReporte(lista):
    """
    funcionamiento: itera sobre una lista ejecutando la función que traduce las sedes y devuelve un string con las sedes traducidas
    entradas: lista: la lista con las sedes
    salidas: el string con las sedes
    """
    salida = ""
    for i in lista:
        salida+= traducirSede(i) +" | "
    return salida[:-2]

def traducirDonador(donador):
    """
    funcionamiento: dependiendo del valor booleano, retorna si o no para donador
    entradas: donador: el valor booelano
    salidas: Si: en caso de que donador sea True
    No: en caso de que el valor booleano sea False
    """
    if donador == True:
        return 'Si'
    else:
        return 'No'
#----------------------------------------------------------------------------#
#                           Base de datos                                    #
#----------------------------------------------------------------------------#
def insertarLicencia(objeto,lista):
    """
    funcionamiento: inserta el objeto generado a la lista
    entradas: objeto: el objeto de la clase Licencia
    lista: la lista con los objetos
    salidas: el objeto licencia con los datos
    """
    lista.append(objeto)
    return lista

def renovarLicencia(posicion, lista):
    """
    Funcionamiento: Se encarga actuallizar a un donador.
    Entradas: -datos: datos del donador matriz: matriz en el que se va a guardar
    Salidas: NA
    """
    fecha = lista[posicion].getFechaVenci()
    annos = convertirFechaAnnos(lista[posicion].getFechaNac())
    if annos >= 18 and annos <= 25:
        fecha = fecha[0:6] + str(int(fecha[6:])+3)
    else:
        fecha = fecha[0:6] + str(int(fecha[6:])+5)
    lista[posicion].setFechaVenci(fecha)


#----------------------------------------------------------------------------#
#                           Reportes                                         #
#----------------------------------------------------------------------------#

def expandirCelda(nombreArchivo):
    """
    Funcionamiento: Expandir las celdas del archivo exel.
    Entradas: nombreArchivo: es el nombre del archivo
    Salidas: Archivo exel
    """
    wb = openpyxl.load_workbook(nombreArchivo)
    worksheet = wb.active
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width
    wb.save(nombreArchivo)

#Procesamiento
def sacarTipoLicencia(tipo, lista):
    """
    Funcionamiento: crea una lista con solo las licencias que son del tipo especificado
    Entradas: -tipo: el tipo de licencia que se busca
    -lista: la lista con todas la licencias
    Salidas: nLista: la lista con las licencias que son del tipo especificado
    """
    nLista = []
    for i in lista:
        if i.getTipoLicencia()[0] == tipo:
            nLista.append(i)
    return nLista

def sacarExamenPorSancion(lista):
    """
    Funcionamiento: crea una lista con solo las licencias que son del tipo especificado
    Entradas: -tipo: el tipo de licencia que se busca
    -lista: la lista con todas la licencias
    Salidas: nLista: la lista con las licencias que son del tipo especificado
    """
    nLista = []
    for i in lista:
        if i.getPuntaje() <= 6 and i.getPuntaje() > 0:
            nLista.append(i)
    return nLista

def sacarDonatesOrganos(lista):
    """
    Funcionamiento: crea una lista con solo las licencias que son del tipo especificado
    Entradas: lista: la lista con todas la licencias
    Salidas: nLista: la lista con las licencias que son del tipo especificado (donantes de Organos)
    """
    nLista = []
    for i in lista:
        if i.getDonador():
            nLista.append(i)
    return nLista

def sacarLicenciasAnuladas(lista):
    """
    Funcionamiento: crea una lista con solo las licencias que son del tipo especificado
    Entradas: lista: la lista con todas la licencias
    Salidas: nLista: la lista con las licencias que son del tipo especificado (Licencias anuladas)
    """
    nLista = []
    for i in lista:
        if i.getPuntaje() == 0:
            nLista.append(i)
    return nLista

def sacarPorSede(lista, sede):
    """
    Funcionamiento: crea una lista con solo las licencias que son del tipo especificado
    Entradas: lista: la lista con todas la licencias
    Salidas: nLista: la lista con las licencias que son del tipo especificado (Sede)
    """
    nLista = []
    for i in lista:
        for j in i.getSede():
            if j == sede:
                nLista.append(i)
    return nLista

#creacion de archivos
def reporteFichaLarga(nombreArch,lista, titulo):
    """
    Funcionamiento: crea el archivo excel con la ficha mas larga que incluye todos los datos de una licencia
    Entradas: nombreArch: el nombre con el que se va a guardar el archivo
    Salidas: NA
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = titulo
    sheet['A2'] = generarFechaExp()+"  "+ obtenerHoraActual()
    sheet['A3'] = 'Cédula'
    sheet['B3'] = 'Nombre'
    sheet['C3'] = 'FechaNac'
    sheet['D3'] = 'FechaExp'
    sheet['E3'] = 'FechaVenc'
    sheet['F3'] = 'TipoLicen'
    sheet['G3'] = 'TipoSangre'
    sheet['H3'] = 'Donador'
    sheet['I3'] = 'Sede'
    sheet['J3'] = 'Puntaje'
    cont = 4
    for i in lista:
        sheet['A'+str(cont)] = i.getCedula()
        sheet['B'+str(cont)] = i.getNombre()
        sheet['C'+str(cont)] = i.getFechaNac()
        sheet['D'+str(cont)] = i.getFechaEx()
        sheet['E'+str(cont)] = i.getFechaVenci()
        sheet['F'+str(cont)] = i.getTipoLicencia()
        sheet['G'+str(cont)] = i.getTipoSangre()
        sheet['H'+str(cont)] = traducirDonador(i.getDonador())
        sheet['I'+str(cont)] = traducirSedeReporte(i.getSede())
        sheet['J'+str(cont)] = i.getPuntaje()
        cont+=1
    workbook.save("ReportesExcel/"+nombreArch)
    expandirCelda("ReportesExcel/"+nombreArch)
    
def reporteFichaCorta(nombreArch,lista,titulo):
    """
    Funcionamiento: crea el archivo excel con la ficha mas corta que incluye todos los datos de una licencia
    Entradas: nombreArch: el nombre con el que se va a guardar el archivo
    Salidas: NA
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = titulo
    sheet['A2'] = generarFechaExp()+"  "+ obtenerHoraActual()
    sheet['A3'] = 'Cédula'
    sheet['B3'] = 'Nombre'
    sheet['C3'] = 'TipoLicen'
    cont = 4
    for i in lista:
        sheet['A'+str(cont)] = i.getCedula()
        sheet['B'+str(cont)] = i.getNombre()
        sheet['C'+str(cont)] = i.getTipoLicencia()
        cont+=1
    workbook.save("ReportesExcel/"+nombreArch)
    expandirCelda("ReportesExcel/"+nombreArch)

def reporteFichaDeCuatro(nombreArch,lista,titulo):
    """
    Funcionamiento: crea el archivo excel con la ficha mas corta que incluye todos los datos de una licencia
    Entradas: nombreArch: el nombre con el que se va a guardar el archivo
    Salidas: NA
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = titulo
    sheet['A2'] = generarFechaExp()+"  "+ obtenerHoraActual()
    sheet['A3'] = 'Cédula'
    sheet['B3'] = 'Nombre'
    sheet['C3'] = 'TipoLicen'
    sheet['D3'] = 'Puntaje'
    cont = 4
    for i in lista:
        sheet['A'+str(cont)] = i.getCedula()
        sheet['B'+str(cont)] = i.getNombre()
        sheet['C'+str(cont)] = i.getTipoLicencia()
        sheet['D'+str(cont)] = i.getPuntaje()
        cont+=1
    workbook.save("ReportesExcel/"+nombreArch)
    expandirCelda("ReportesExcel/"+nombreArch)

def reporteFichaPuntaje(nombreArch,lista,titulo):
    """
    Funcionamiento: crea el archivo excel con la ficha sin puntaje
    Entradas: nombreArch: el nombre con el que se va a guardar el archivo
    Salidas: NA
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = titulo
    sheet['A2'] = generarFechaExp()+"  "+ obtenerHoraActual()
    sheet['A3'] = 'Cédula'
    sheet['B3'] = 'Nombre'
    sheet['C3'] = 'FechaNac'
    sheet['D3'] = 'FechaExp'
    sheet['E3'] = 'FechaVenc'
    sheet['F3'] = 'TipoLicen'
    sheet['G3'] = 'TipoSangre'
    sheet['H3'] = 'Donador'
    sheet['I3'] = 'Sede'
    cont = 4
    for i in lista:
        sheet['A'+str(cont)] = i.getCedula()
        sheet['B'+str(cont)] = i.getNombre()
        sheet['C'+str(cont)] = i.getFechaNac()
        sheet['D'+str(cont)] = i.getFechaEx()
        sheet['E'+str(cont)] = i.getFechaVenci()
        sheet['F'+str(cont)] = i.getTipoLicencia()
        sheet['G'+str(cont)] = i.getTipoSangre()
        sheet['H'+str(cont)] = traducirDonador(i.getDonador())
        sheet['I'+str(cont)] = traducirSedeReporte(i.getSede())
        cont+=1
    workbook.save("ReportesExcel/"+nombreArch)
    expandirCelda("ReportesExcel/"+nombreArch)